from flask import Flask, render_template, request, jsonify, send_file
import re
from collections import defaultdict
import os
import requests
import json
from datetime import datetime
import tempfile
import uuid
import traceback

# Import PDF library directly
try:
    from pypdf import PdfReader
except ImportError:
    try:
        from PyPDF2 import PdfReader
    except ImportError:
        PdfReader = None
        print("Warning: Neither pypdf nor PyPDF2 is available")

def _import_openpyxl():
    global Workbook, Font, PatternFill, Alignment, Border, Side, get_column_letter
    if Workbook is None:
        try:
            from openpyxl import Workbook as _Workbook
            from openpyxl.styles import Font as _Font, PatternFill as _PatternFill, Alignment as _Alignment, Border as _Border, Side as _Side
            from openpyxl.utils import get_column_letter as _get_column_letter
            Workbook = _Workbook
            Font = _Font
            PatternFill = _PatternFill
            Alignment = _Alignment
            Border = _Border
            Side = _Side
            get_column_letter = _get_column_letter
        except Exception as e:
            print(f"Warning: openpyxl import failed: {e}")
    return Workbook

# Determine if running on Vercel
IS_VERCEL = os.environ.get('VERCEL') == '1'

# Initialize Flask app with explicit template folder
# In Vercel, we need to ensure templates are found
try:
    base_path = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(base_path, 'templates')
    # Check if templates directory exists
    if os.path.exists(template_path):
        app = Flask(__name__, template_folder=template_path)
        print(f"Flask initialized with template folder: {template_path}")
    else:
        # Use default (Flask will look for 'templates' directory)
        app = Flask(__name__)
        print(f"Flask initialized with default template folder (templates directory not found at {template_path})")
except Exception as e:
    print(f"Error initializing Flask: {e}")
    # Fallback to default
    app = Flask(__name__)

# For Vercel/Render, use /tmp for uploads (serverless-friendly)
# For local development, use 'uploads'
# Check if we're in a cloud environment (Vercel or Render)
IS_CLOUD = os.environ.get('VERCEL') == '1' or os.environ.get('RENDER') == 'true'
UPLOAD_FOLDER = '/tmp/uploads' if IS_CLOUD else 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Create uploads directory if it doesn't exist
try:
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    print(f"Upload folder created/verified: {app.config['UPLOAD_FOLDER']}")
except Exception as e:
    print(f"Warning: Could not create upload folder: {e}")
    # Try to use /tmp as fallback
    try:
        app.config['UPLOAD_FOLDER'] = '/tmp'
        print(f"Using fallback upload folder: /tmp")
    except:
        pass

# Claude API configuration
CLAUDE_API_URL = 'https://api.anthropic.com/v1/messages'

def get_claude_api_key():
    """Get Claude API key, raising error if not set"""
    api_key = os.environ.get('CLAUDE_API_KEY')
    if not api_key:
        raise ValueError("CLAUDE_API_KEY environment variable is required. Please set it in Vercel dashboard: Settings > Environment Variables")
    return api_key

# Expense categories
CATEGORIES = [
    'Food & Dining',
    'Transportation',
    'Shopping',
    'Bills & Utilities',
    'Entertainment',
    'Healthcare',
    'Personal Care',
    'Education',
    'Other'
]

def extract_expenses_from_pdf(pdf_path):
    """Extract expense transactions from PDF statement using text extraction"""
    if PdfReader is None:
        raise ImportError("pypdf is not available. Please install it with: pip install pypdf")
    
    expenses = []
    
    try:
        print(f"Opening PDF: {pdf_path}")
        with open(pdf_path, 'rb') as file:
            pdf_reader = PdfReader(file)
            num_pages = len(pdf_reader.pages)
            print(f"PDF has {num_pages} pages")
            
            # Extract text from all pages
            full_text = ""
            for i, page in enumerate(pdf_reader.pages):
                if i % 10 == 0:
                    print(f"Extracting text from page {i+1}/{num_pages}...")
                full_text += page.extract_text() + "\n"
            
            print(f"Extracted {len(full_text)} characters of text")
            
            # Pattern to match transaction lines
            # Format: DD-MMM-YY Description Amount (AED optional)
            # Example: "08-Oct-24 NFC - (AP-PAY)-DUBAI MALL AED 150.00"
            transaction_pattern = re.compile(
                r'(\d{1,2}-[A-Za-z]{3}-\d{2})\s+'  # Date: DD-MMM-YY
                r'(.+?)'  # Description (non-greedy)
                r'\s+(?:AED\s+)?([\d,]+\.?\d*)\s*'  # Amount (optional AED prefix)
                r'(?:\n|$)',  # End of line
                re.MULTILINE | re.IGNORECASE
            )
            
            # Alternative pattern for lines without AED prefix
            transaction_pattern_alt = re.compile(
                r'(\d{1,2}-[A-Za-z]{3}-\d{2})\s+'  # Date
                r'(.+?)'  # Description
                r'\s+([\d,]+\.?\d{2})\s*$',  # Amount with 2 decimals
                re.MULTILINE | re.IGNORECASE
            )
            
            # Find all transaction matches
            matches = transaction_pattern.findall(full_text)
            matches_alt = transaction_pattern_alt.findall(full_text)
            
            # Combine matches, preferring the first pattern
            all_matches = matches + [m for m in matches_alt if m not in matches]
            
            seen_transactions = set()
            
            for match in all_matches:
                if len(match) < 3:
                    continue
                
                posting_date = match[0].strip()
                description = match[1].strip()
                amount_str = match[2].strip()
                
                # Skip empty or invalid entries
                if not posting_date or not description or not amount_str:
                    continue
                
                # Skip header-like rows
                skip_keywords = ['opening balance', 'closing balance', 'total outstanding', 
                               'transaction date', 'posting date', 'transaction details', 
                               'original amount', 'total amount', 'important:', 'warning',
                               'page', 'statement', 'account']
                description_lower = description.lower()
                if any(keyword in description_lower for keyword in skip_keywords):
                    continue
                
                # Skip credits (amounts with CR or negative)
                if ' CR' in amount_str.upper() or amount_str.upper().strip().endswith('CR'):
                    continue
                
                # Validate date format
                if not re.match(r'\d{1,2}-[A-Za-z]{3}-\d{2}', posting_date):
                    continue
                
                # Clean and parse amount
                amount_clean = re.sub(r'[^\d.,-]', '', amount_str)
                amount_clean = amount_clean.replace(',', '').strip()
                
                if not amount_clean:
                    continue
                
                try:
                    amount_float = float(amount_clean)
                    if amount_float <= 0:
                        continue
                    
                    # Clean description
                    description_clean = ' '.join(description.split())
                    
                    # Skip if description is too short
                    if len(description_clean) < 3:
                        continue
                    
                    # Create transaction key to avoid duplicates
                    transaction_key = (posting_date, description_clean, amount_float)
                    if transaction_key in seen_transactions:
                        continue
                    seen_transactions.add(transaction_key)
                    
                    expenses.append({
                        'date': posting_date,
                        'description': description_clean,
                        'amount': amount_float
                    })
                except (ValueError, AttributeError):
                    continue
            
            # If no transactions found with regex, try line-by-line parsing
            if not expenses:
                lines = full_text.split('\n')
                current_date = None
                
                for line in lines:
                    line = line.strip()
                    if not line:
                        continue
                    
                    # Check if line starts with a date
                    date_match = re.match(r'(\d{1,2}-[A-Za-z]{3}-\d{2})', line)
                    if date_match:
                        current_date = date_match.group(1)
                        # Try to extract amount from same line
                        amount_match = re.search(r'([\d,]+\.?\d{2})', line)
                        if amount_match:
                            desc_part = line[len(date_match.group(0)):amount_match.start()].strip()
                            amount_str = amount_match.group(1)
                            if desc_part and len(desc_part) > 3:
                                try:
                                    amount_float = float(amount_str.replace(',', ''))
                                    if amount_float > 0 and 'CR' not in line.upper():
                                        expenses.append({
                                            'date': current_date,
                                            'description': desc_part,
                                            'amount': amount_float
                                        })
                                except ValueError:
                                    pass
    
    except Exception as e:
        print(f"Error extracting from PDF: {e}")
        raise
    
    return expenses

def clean_description(description):
    """Clean transaction description"""
    # Remove leading "- " and extra whitespace
    desc = description.strip()
    if desc.startswith('- '):
        desc = desc[2:].strip()
    # Remove NFC/IAP prefixes that don't add meaning
    desc = re.sub(r'^(NFC|IAP)\s*-\s*\([^)]+\)\s*-\s*', '', desc, flags=re.IGNORECASE)
    return desc.strip()

def categorize_expense_with_claude(description):
    """Categorize an expense using Claude API"""
    cleaned_desc = clean_description(description)
    categories_str = ', '.join(CATEGORIES)
    
    prompt = f"""Categorize this bank transaction into ONE of these categories:
{categories_str}

Transaction: "{cleaned_desc}"

Examples:
- "Amazon.ae Dubai" or "Amazon Now" → Shopping
- "Carrefour" or "Lulu" or supermarket → Food & Dining
- "Restaurant" or "Food" or "Cafe" → Food & Dining
- "Uber" or "Careem" or "Taxi" or "Metro" → Transportation
- "Dubai Government" or "DEWA" or "Etisalat" → Bills & Utilities
- "Sephora" or "Nike" → Shopping
- "Hospital" or "Pharmacy" → Healthcare

Respond with ONLY the exact category name from the list above, nothing else."""

    try:
        api_key = get_claude_api_key()
        headers = {
            'x-api-key': api_key,
            'anthropic-version': '2023-06-01',
            'Content-Type': 'application/json'
        }
        
        data = {
            'model': 'claude-3-haiku-20240307',
            'max_tokens': 50,
            'temperature': 0.1,
            'messages': [{
                'role': 'user',
                'content': [{
                    'type': 'text',
                    'text': prompt
                }]
            }]
        }
        
        response = requests.post(CLAUDE_API_URL, headers=headers, json=data, timeout=10)
        response.raise_for_status()
        
        result = response.json()
        category = result['content'][0]['text'].strip()
        
        # Validate category is in our list
        if category in CATEGORIES:
            return category
        else:
            # Try to find a match (case-insensitive, partial)
            for cat in CATEGORIES:
                if cat.lower() == category.lower() or cat.lower() in category.lower() or category.lower() in cat.lower():
                    return cat
            return 'Other'
    except Exception as e:
        print(f"Error categorizing with Claude: {e}")
        return 'Other'

def categorize_expenses_batch(expenses, batch_size=50):
    """Categorize multiple expenses using Claude API in smaller batches"""
    if not expenses:
        return {}
    
    # For very large statements, use a faster approach
    if len(expenses) > 200:
        print(f"Large statement detected ({len(expenses)} transactions). Using optimized processing...")
        batch_size = 50  # Larger batches for speed
    
    all_categorized = {}
    total_batches = (len(expenses) + batch_size - 1) // batch_size
    
    # Process in smaller batches to avoid token limits and improve accuracy
    for batch_num, batch_start in enumerate(range(0, len(expenses), batch_size), 1):
        print(f"Processing batch {batch_num}/{total_batches}...")
        batch_expenses = expenses[batch_start:batch_start + batch_size]
        batch_indices = list(range(batch_start, batch_start + len(batch_expenses)))
        
        # Clean descriptions
        transactions_text = "\n".join([
            f"{i+1}. {clean_description(exp['description'])}" 
            for i, exp in enumerate(batch_expenses)
        ])
        categories_str = ', '.join(CATEGORIES)
        
        prompt = f"""Categorize these UAE bank transactions into ONE category each from this list:
{categories_str}

Transactions:
{transactions_text}

Guidelines:
- Amazon, online shopping, retail stores → Shopping
- Carrefour, Lulu, supermarkets, grocery stores, food markets → Food & Dining
- Restaurants, cafes, food delivery, fast food → Food & Dining
- Uber, Careem, taxi, metro, transportation → Transportation
- Government services, utilities, phone/internet bills → Bills & Utilities
- Pharmacies, hospitals, medical → Healthcare
- Salons, gyms, fitness → Personal Care

Respond with a JSON object: {{"1": "Category Name", "2": "Category Name", ...}}
Use exact category names from the list. Only respond with valid JSON, no other text."""

        try:
            api_key = get_claude_api_key()
            headers = {
                'x-api-key': api_key,
                'anthropic-version': '2023-06-01',
                'Content-Type': 'application/json'
            }
            
            data = {
                'model': 'claude-3-haiku-20240307',
                'max_tokens': 1000,
                'temperature': 0.1,
                'messages': [{
                    'role': 'user',
                    'content': [{
                        'type': 'text',
                        'text': prompt
                    }]
                }]
            }
            
            response = requests.post(CLAUDE_API_URL, headers=headers, json=data, timeout=60)
            response.raise_for_status()
            
            result = response.json()
            response_text = result['content'][0]['text'].strip()
            
            # Try to extract JSON from response
            try:
                # Remove markdown code blocks if present
                response_text = re.sub(r'```json\s*', '', response_text)
                response_text = re.sub(r'```\s*', '', response_text)
                # Extract JSON object (handle nested objects)
                json_match = re.search(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', response_text)
                if json_match:
                    response_text = json_match.group(0)
                else:
                    # Try to find JSON starting from first {
                    start_idx = response_text.find('{')
                    if start_idx >= 0:
                        # Find matching closing brace
                        brace_count = 0
                        end_idx = start_idx
                        for i in range(start_idx, len(response_text)):
                            if response_text[i] == '{':
                                brace_count += 1
                            elif response_text[i] == '}':
                                brace_count -= 1
                                if brace_count == 0:
                                    end_idx = i + 1
                                    break
                        if end_idx > start_idx:
                            response_text = response_text[start_idx:end_idx]
                
                categories_dict = json.loads(response_text)
                
                # Map transaction indices to categories
                for i, expense_idx in enumerate(batch_indices):
                    key = str(i + 1)
                    category = categories_dict.get(key, 'Other')
                    # Validate category
                    if category not in CATEGORIES:
                        # Try fuzzy match
                        category_lower = category.lower()
                        matched = False
                        for cat in CATEGORIES:
                            if cat.lower() == category_lower or cat.lower() in category_lower:
                                category = cat
                                matched = True
                                break
                        if not matched:
                            category = 'Other'
                    all_categorized[expense_idx] = category
            except (json.JSONDecodeError, KeyError) as e:
                print(f"Error parsing batch response: {e}")
                # Fallback: categorize individually for this batch
                for expense_idx in batch_indices:
                    all_categorized[expense_idx] = categorize_expense_with_claude(expenses[expense_idx]['description'])
                
        except Exception as e:
            print(f"Error categorizing batch with Claude: {e}")
            # Fallback: categorize individually for this batch
            for expense_idx in batch_indices:
                all_categorized[expense_idx] = categorize_expense_with_claude(expenses[expense_idx]['description'])
    
    return all_categorized

def process_statement(pdf_path):
    """Process PDF statement and return categorized expenses"""
    print(f"Starting to process statement: {pdf_path}")
    
    expenses = extract_expenses_from_pdf(pdf_path)
    print(f"Extracted {len(expenses)} expenses from PDF")
    
    if not expenses:
        return {
            'categories': {},
            'total_expenses': 0.0,
            'total_transactions': 0
        }
    
    # Limit processing to first 300 transactions to avoid Render timeout (30s limit)
    if len(expenses) > 300:
        print(f"Large statement ({len(expenses)} transactions). Processing first 300 only to avoid timeout.")
        expenses = expenses[:300]
        
        # For very large statements, use keyword-based categorization first
        # Then use Claude API only for uncategorized transactions
        if len(expenses) > 200:
            print(f"Large statement ({len(expenses)} transactions). Using hybrid approach...")
        categorized = defaultdict(lambda: {'total': 0.0, 'transactions': []})
        uncategorized = []
        uncategorized_indices = []
        
        # Quick keyword-based categorization
        keyword_categories = {
            'food': 'Food & Dining',
            'restaurant': 'Food & Dining',
            'cafe': 'Food & Dining',
            'carrefour': 'Food & Dining',
            'lulu': 'Food & Dining',
            'supermarket': 'Food & Dining',
            'grocery': 'Food & Dining',
            'uber': 'Transportation',
            'careem': 'Transportation',
            'taxi': 'Transportation',
            'metro': 'Transportation',
            'amazon': 'Shopping',
            'shopping': 'Shopping',
            'mall': 'Shopping',
            'pharmacy': 'Healthcare',
            'hospital': 'Healthcare',
            'medical': 'Healthcare',
            'salon': 'Personal Care',
            'gym': 'Personal Care',
            'fitness': 'Personal Care',
            'etisalat': 'Bills & Utilities',
            'du': 'Bills & Utilities',
            'dewa': 'Bills & Utilities',
            'utility': 'Bills & Utilities',
        }
        
        for i, expense in enumerate(expenses):
            desc_lower = expense['description'].lower()
            categorized_flag = False
            
            for keyword, category in keyword_categories.items():
                if keyword in desc_lower:
                    categorized[category]['total'] += expense['amount']
                    categorized[category]['transactions'].append(expense)
                    categorized_flag = True
                    break
            
            if not categorized_flag:
                uncategorized.append(expense)
                uncategorized_indices.append(i)
        
        print(f"Keyword categorization: {len(expenses) - len(uncategorized)} categorized, {len(uncategorized)} need API")
        
        # Use Claude API only for uncategorized transactions
        if uncategorized:
            batch_categories = categorize_expenses_batch(uncategorized, batch_size=50)
            for i, expense in enumerate(uncategorized):
                category = batch_categories.get(i, 'Other')
                categorized[category]['total'] += expense['amount']
                categorized[category]['transactions'].append(expense)
    else:
        # Try batch categorization first (more efficient)
        print(f"Using batch categorization for {len(expenses)} transactions...")
        batch_categories = categorize_expenses_batch(expenses)
    
        categorized = defaultdict(lambda: {'total': 0.0, 'transactions': []})
        
        if batch_categories:
            # Use batch results
            for i, expense in enumerate(expenses):
                category = batch_categories.get(i, 'Other')
                categorized[category]['total'] += expense['amount']
                categorized[category]['transactions'].append(expense)
        else:
            # Fallback to individual categorization
            print("Falling back to individual categorization...")
            for i, expense in enumerate(expenses):
                if i % 50 == 0:
                    print(f"Processing transaction {i+1}/{len(expenses)}...")
                category = categorize_expense_with_claude(expense['description'])
                categorized[category]['total'] += expense['amount']
                categorized[category]['transactions'].append(expense)
    
    # Convert to regular dict and sort by total
    result = {
        'categories': dict(sorted(categorized.items(), key=lambda x: x[1]['total'], reverse=True)),
        'total_expenses': sum(cat['total'] for cat in categorized.values()),
        'total_transactions': len(expenses)
    }
    
    print(f"Processing complete. Returning {len(result['categories'])} categories")
    return result

@app.route('/')
def index():
    try:
        # Log template path for debugging
        template_path = app.template_folder or 'templates'
        print(f"Template folder: {template_path}")
        print(f"Looking for index.html in: {os.path.join(template_path, 'index.html')}")
        return render_template('index.html')
    except Exception as e:
        import traceback
        error_msg = f"Error loading template: {str(e)}\n{traceback.format_exc()}"
        print(error_msg)
        # Return a simple HTML error page with full error details
        return f"""
        <html>
        <head><title>Error</title></head>
        <body>
        <h1>Error Loading Page</h1>
        <p><strong>Error:</strong> {str(e)}</p>
        <pre>{traceback.format_exc()}</pre>
        <p>Template folder: {app.template_folder or 'default'}</p>
        </body>
        </html>
        """, 500

@app.route('/favicon.ico')
def favicon():
    """Return empty favicon to avoid 500 errors"""
    return '', 204

@app.route('/test')
def test():
    """Simple test endpoint that doesn't use templates"""
    return jsonify({'message': 'Flask is working!', 'status': 'ok'}), 200

@app.route('/health')
def health():
    """Health check endpoint"""
    try:
        return jsonify({
            'status': 'ok',
            'pypdf': pypdf is not None,
            'openpyxl': Workbook is not None,
            'vercel': IS_VERCEL,
            'claude_api_key_set': bool(os.environ.get('CLAUDE_API_KEY')),
            'template_folder': str(app.template_folder) if app.template_folder else 'default'
        }), 200
    except Exception as e:
        import traceback
        return jsonify({
            'status': 'error', 
            'message': str(e),
            'traceback': traceback.format_exc()
        }), 500

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'Please upload a PDF file'}), 400
    
    filepath = None
    try:
        # Ensure upload directory exists
        upload_dir = app.config['UPLOAD_FOLDER']
        os.makedirs(upload_dir, exist_ok=True)
        
        # Use a unique filename to avoid conflicts
        unique_filename = f"{uuid.uuid4()}_{file.filename}"
        filepath = os.path.join(upload_dir, unique_filename)
        
        # Save uploaded file
        file.save(filepath)
        print(f"File saved to: {filepath}")
        
        # Process the statement
        result = process_statement(filepath)
        
        # Clean up uploaded file
        if filepath and os.path.exists(filepath):
            os.remove(filepath)
        
        return jsonify(result)
    
    except Exception as e:
        error_msg = str(e)
        error_trace = traceback.format_exc()
        print(f"Upload error: {error_msg}")
        print(f"Traceback: {error_trace}")
        
        # Clean up uploaded file if it exists
        if filepath and os.path.exists(filepath):
            try:
                os.remove(filepath)
            except:
                pass
        
        return jsonify({
            'error': f'Error processing file: {error_msg}',
            'details': error_trace if app.debug else None
        }), 500

@app.route('/export', methods=['POST'])
def export_to_excel():
    """Export categorized expenses to Excel"""
    Workbook_module = _import_openpyxl()
    if Workbook_module is None or Font is None:
        return jsonify({'error': 'Excel export not available: openpyxl not installed'}), 500
    
    try:
        data = request.json
        if not data or 'categories' not in data:
            return jsonify({'error': 'No data to export'}), 400
        
        # Create a new workbook
        wb = Workbook_module()
        ws = wb.active
        ws.title = "Expense Report"
        
        # Define styles (using imported classes)
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        category_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        category_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        currency_format = '#,##0.00'
        
        # Header row
        headers = ['Date', 'Description', 'Amount (AED)', 'Category']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        
        # Write data
        row_num = 2
        total_expenses = 0
        
        # Sort categories by total amount
        sorted_categories = sorted(data['categories'].items(), key=lambda x: x[1]['total'], reverse=True)
        
        for category, info in sorted_categories:
            # Category header row
            category_row = row_num
            ws.merge_cells(f'A{category_row}:D{category_row}')
            cell = ws.cell(row=category_row, column=1)
            cell.value = f"{category} - Total: AED {info['total']:.2f}"
            cell.fill = category_fill
            cell.font = category_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border
            row_num += 1
            
            # Transactions in this category
            for transaction in info['transactions']:
                ws.cell(row=row_num, column=1, value=transaction.get('date', '')).border = border
                ws.cell(row=row_num, column=2, value=transaction.get('description', '')).border = border
                amount_cell = ws.cell(row=row_num, column=3, value=transaction.get('amount', 0))
                amount_cell.number_format = currency_format
                amount_cell.border = border
                ws.cell(row=row_num, column=4, value=category).border = border
                
                # Alignments
                ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
                ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='right')
                ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='left')
                
                total_expenses += transaction.get('amount', 0)
                row_num += 1
            
            # Empty row between categories
            row_num += 1
        
        # Summary section
        summary_row = row_num + 1
        ws.merge_cells(f'A{summary_row}:B{summary_row}')
        ws.cell(row=summary_row, column=1, value="TOTAL EXPENSES:").font = Font(bold=True, size=12)
        ws.cell(row=summary_row, column=3, value=total_expenses).number_format = currency_format
        ws.cell(row=summary_row, column=3).font = Font(bold=True, size=12)
        
        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.close()
        wb.save(temp_file.name)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'expense_report_{timestamp}.xlsx'
        
        return send_file(
            temp_file.name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'error': f'Error exporting to Excel: {str(e)}'}), 500

# Vercel handler - must be at the very end of the file
# Vercel Python runtime automatically wraps Flask apps
# Export the Flask app as 'handler' for Vercel
handler = app

if __name__ == '__main__':
    # Get port from environment variable (Railway/Render provide this)
    port = int(os.environ.get('PORT', 5019))
    app.run(host='0.0.0.0', port=port, debug=False)

